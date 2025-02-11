import tkinter as tk
from tkinter import filedialog, messagebox, ttk  # Import ttk for themed widgets
import os
import openpyxl
from openpyxl import load_workbook
import sys
from collections import defaultdict
import threading # Importando a biblioteca de threading

import openpyxl.workbook

# ============================================================
# Script para processamento de Clashs e geração de outputs.
# Este script contém funções para:
# - Selecionar arquivos de entrada (Clash e Matriz)
# - Processar o arquivo de clash e extrair informações
# - Validar e separar os clashs (aprovados, com exceção, problemáticos)
# - Gerar arquivos de texto e planilhas de Excel com os resultados
# - Fornecer uma interface gráfica utilizando Tkinter
# ============================================================

# Variáveis globais para armazenar os caminhos dos arquivos/diretórios
clash_file_path = None
matrix_file_path = None
output_dir = None

# Variável global para rastrear os clashs já contados por disciplina.
# Ela é um dicionário onde cada disciplina mapeia para um conjunto de IDs de clash processados.
clashs_contados = defaultdict(set)

# ------------------------------------------------------------
# Função: select_clash_file
# Propósito: Abre um diálogo para o usuário selecionar o arquivo de clash (.txt)
# e armazena o caminho na variável global clash_file_path.
# ------------------------------------------------------------
def select_clash_file():
    filename = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    if filename:
        clash_file_path.set(filename)

# ------------------------------------------------------------
# Função: select_matrix_file
# Propósito: Abre um diálogo para o usuário selecionar o arquivo da matriz (.xlsx)
# e armazena o caminho na variável global matrix_file_path.
# ------------------------------------------------------------
def select_matrix_file():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filename:
        matrix_file_path.set(filename)

# ------------------------------------------------------------
# Função: select_output_dir
# Propósito: Abre um diálogo para o usuário selecionar o diretório de saída
# e armazena o caminho na variável global output_dir.
# ------------------------------------------------------------
def select_output_dir():
    dirname = filedialog.askdirectory()
    if dirname:
        output_dir.set(dirname)

# ------------------------------------------------------------
# Função: extract_disciplina
# Propósito: Extrai a disciplina a partir de uma linha do caminho (linha que começa com "Path:")
# Utiliza splits sucessivos para isolar a informação e mapeia a sigla da disciplina para seu nome.
# ------------------------------------------------------------
def extract_disciplina(linha):
    """
    Extrai a disciplina a partir de uma linha do caminho (linha que começa com "Path:").
    Realiza uma série de splits conforme o padrão esperado da string e retorna a disciplina
    associada, ou None caso não consiga extrair.
    """
    try:
        # Primeiramente ignora as partes antes dos '>':
        _, rsp = linha.split('>', 1)
        _, rsp2 = rsp.split('>', 1)
        
        # Processa os splits consecutivos conforme o padrão:
        temp = rsp2
        _, temp = temp.split('-', 1)      # descarta a primeira parte (valrsp)
        _, temp = temp.split('-', 1)      # descarta a segunda parte (valrj)
        _, temp = temp.split('-', 1)      # descarta a terceira parte (val218)
        _, temp = temp.split('-', 1)      # descarta a quarta parte (val226)
        _, temp = temp.split('-', 1)      # descarta a quinta parte (valaca)
        _, temp = temp.split('-', 1)      # descarta a sexta parte (valexe)
        _, temp = temp.split('-', 1)      # descarta a sétima parte (valmb), 
        # Agora temp contém "valsemifinal-numdps-resto..."
        valfinal, temp = temp.split('-', 1)
        numdps, _ = temp.split('-', 1)
        
        # Limpa os espaços em branco
        valfinal = valfinal.strip()
        numdps = numdps.strip()
        
        # Mapeamento básico das siglas para as disciplinas
        disciplinas = {
            'C1': 'Topografia',
            'F1': 'Geometria',
            'G1': 'Terraplenagem',
            'H2': 'Drenagem',
            'J2': 'Dispositivos de Segurança',
            'I2': 'Pavimentação',
            'L2': 'OAEs',
            'K2': 'Iluminação',
            'L4': 'Contenções',
            'M1': 'Interferências',
            'Q1': 'Desapropriação',
            'N2': 'Paisagismo',
            'Z9': 'Geral'
        }
        
        # Tratamento especial para 'J1'
        if valfinal == 'J1' and numdps == '001':
            return 'Sinalização Vertical'
        
        return disciplinas.get(valfinal)
    except Exception as e:
        print(f"Erro durante a extração da disciplina: {e}")
        return None

# ------------------------------------------------------------
# Função: is_clash_complete
# Propósito: Verificar se um registro de clash possui todas as informações obrigatórias.
# Retorna True se estiver completo; caso contrário, retorna False.
# ------------------------------------------------------------
def is_clash_complete(clash):
    """
    Verifica se o dicionário do clash contém todas as informações obrigatórias.
    Caso algum campo obrigatório esteja faltando ou com valor vazio (e para os layers,
    se o valor for 'Layer_vazio'), a função retorna False.
    """
    required_fields = [
        'name', 'id', 'coord_x', 'coord_y', 'coord_z',
        'disciplina_1', 'disciplina_2',
        'layer_1', 'layer_2'
    ]
    for field in required_fields:
        if field not in clash or not clash[field]:
            return False
        if field in ['layer_1', 'layer_2'] and clash[field] == 'Layer_vazio':
            return False
    return True

# ------------------------------------------------------------
# Função: process_clash_file
# Propósito: Processa o arquivo de clash (.txt), separando registros completos dos problemáticos.
# Retorna uma lista de clashs completos, a lista de disciplinas encontradas e os clashs problemáticos.
# ------------------------------------------------------------
def process_clash_file(filepath):
    lista_disciplinas = []
    clashs = []               # Registros completos
    clashs_problematicos = [] # Registros com informações faltantes
    current_clash = {}
    
    with open(filepath, 'r', encoding='utf-8') as arquivo:
        linhas = arquivo.readlines()
    
    for i in range(len(linhas)):
        linha = linhas[i].strip()
        
        if linha.startswith('Name:'):
            if current_clash:  # Se já houver um clash em progresso, verifica se está completo
                if is_clash_complete(current_clash):
                    clashs.append(current_clash)
                else:
                    clashs_problematicos.append(current_clash)
            current_clash = {}  # Inicia um novo registro de clash
            _, valor = linha.split(':', 1)
            current_clash['name'] = valor.strip()
        
        elif linha.startswith('Image Location:'):
            _, valor = linha.split(':', 1)
            _, pre_id = linha.split('\\', 1)
            id_val, _ = pre_id.split('.', 1)
            current_clash['image_loc'] = valor.strip()
            current_clash['id'] = id_val.strip()

        elif linha.startswith('Clash Point:'):
            _, valor = linha.split(':', 1)
            cord_sem_m = valor.replace('m', '').strip()
            x, resto = cord_sem_m.split(',', 1)
            y, z = resto.split(',', 1)
            current_clash['coord_x'] = x.strip()
            current_clash['coord_y'] = y.strip()
            current_clash['coord_z'] = z.strip()
            current_clash['coordinates'] = cord_sem_m

        elif linha.startswith('Path:'):
            disciplina = extract_disciplina(linha)
            if disciplina:
                if disciplina not in lista_disciplinas:
                    lista_disciplinas.append(disciplina)
                key = 'disciplina_1' if 'disciplina_1' not in current_clash else 'disciplina_2'
                current_clash[key] = disciplina

        elif linha.startswith('Entity Handle:'):
            if 'entity_1' not in current_clash:
                _, valor = linha.split(':', 1)
                current_clash['entity_1'] = valor.strip()
            else:
                _, valor = linha.split(':', 1)
                current_clash['entity_2'] = valor.strip()
        
        elif linha.startswith('Item 1'):
            if linhas[i+1].startswith('Layer:'):
                _, valor = linhas[i+1].split(':', 1)
                current_clash['layer_1'] = valor.strip()
            else:
                current_clash['layer_1'] = 'Layer_vazio'

        elif linha.startswith('Item 2'):
            if linhas[i+1].startswith('Layer:'):
                _, valor = linhas[i+1].split(':', 1)
                current_clash['layer_2'] = valor.strip()
            else:
                current_clash['layer_2'] = 'Layer_vazio'

    # Verifica o último clash processado após o loop
    if current_clash:
        if is_clash_complete(current_clash):
            clashs.append(current_clash)
        else:
            clashs_problematicos.append(current_clash)
    return clashs, lista_disciplinas, clashs_problematicos

# ------------------------------------------------------------
# Função: process_matrix
# Propósito: Processar a planilha de matriz (.xlsx) para validar os clashs com base nas disciplinas.
# Retorna a lista de clashs semi aprovados conforme verificação na matriz.
# ------------------------------------------------------------
def process_matrix(clashs, matrix_path):
    workbook = openpyxl.load_workbook(matrix_path)
    aba_matriz = workbook['Matriz']

    # Inicializa a lista que armazenará os clashs aprovados pela matriz
    clashs_semi_aprovados = []

    for clash in clashs:
        coluna = None  # Coluna correspondente à disciplina_1
        row = None     # Linha correspondente à disciplina_2
        
        # Busca a coluna que contém o valor de 'disciplina_1' na segunda linha (colunas a partir da 3)
        for col in aba_matriz.iter_cols(min_row=2, max_row=2, min_col=3):
            for cel in col:
                if cel.value == clash['disciplina_1']:
                    coluna = cel.column
                    break
            if coluna is not None:
                break
        
        # Busca a linha que contém 'disciplina_2' na coluna 2 (linhas a partir da 3)
        for col in aba_matriz.iter_cols(min_col=2, max_col=2, min_row=3):
            for cel in col:
                if cel.value == clash['disciplina_2']:
                    row = cel.row
                    break
            if row is not None:
                break
                
        # Se o clash envolver "Topografia", imprime detalhes para debug
        '''
        if "Topografia" in (clash.get('disciplina_1', ''), clash.get('disciplina_2', '')):
            print("DEBUG: Clash de Topografia encontrado:")
            print(f"   ID: {clash.get('id', 'N/A')}")
            print(f"   Disciplina 1: {clash.get('disciplina_1', 'N/A')}, Disciplina 2: {clash.get('disciplina_2', 'N/A')}")
            print(f"   Coluna definida para disciplina_1: {coluna} | Linha definida para disciplina_2: {row}")
            if coluna is not None and row is not None:
                cell_value = aba_matriz.cell(row=row, column=coluna).value
                print(f"   Valor da célula na matriz: {cell_value}")
            else:
                print("   Não foram encontradas coluna ou linha correspondentes para esse clash.")
        '''
        # Se ambos 'coluna' e 'row' forem encontrados e a célula correspondente for "O",
        # o clash é adicionado à lista de semi aprovados.
        if coluna is not None and row is not None:
            if aba_matriz.cell(row=row, column=coluna).value == 'O':
                clashs_semi_aprovados.append(clash)
    return clashs_semi_aprovados

# ------------------------------------------------------------
# Função: separacao_de_excecao
# Propósito: Verificar quais clashs aprovados pela matriz estão na lista de exceções
# definida na aba 'exceções' e separá-los em duas listas (aprovados e exceções).
# ------------------------------------------------------------
def separacao_de_excecao(clashs, matriz_path):
    planilha = openpyxl.load_workbook(matriz_path)
    aba = planilha['exceções']
    clashs_aprovados = []
    clashs_excecoes = []
    for clash in clashs:
        # Obtém e limpa os layers
        layer1 = str(clash.get('layer_1', '')).strip()
        layer2 = str(clash.get('layer_2', '')).strip()
        is_excecao = False  # Flag para identificar se o clash está na lista de exceções
        
        # Verifica cada linha na aba de exceções
        for i in range(2, aba.max_row + 1):
            cell_ex1 = aba.cell(row=i, column=2)
            cell_ex2 = aba.cell(row=i, column=4)
            layer_ex1 = str(cell_ex1.value).strip() if cell_ex1.value is not None else ''
            layer_ex2 = str(cell_ex2.value).strip() if cell_ex2.value is not None else ''
            
            if (layer1 == layer_ex1 and layer2 == layer_ex2) or (layer1 == layer_ex2 and layer2 == layer_ex1):
                is_excecao = True
                break
        
        if not is_excecao:
            clashs_aprovados.append(clash)
        elif is_excecao:
            clashs_excecoes.append(clash)
    
    return clashs_aprovados, clashs_excecoes

# ------------------------------------------------------------
# Função: criar_txts_por_disciplina
# Propósito: Para cada registro de clash, cria/atualiza arquivos de texto separados
# por disciplina, caso o clash ainda não tenha sido contado para aquela disciplina.
# ------------------------------------------------------------
def criar_txts_por_disciplina(clashs_total, diretorio_saida):
    # Cria o diretório de saída, se ele não existir.
    # É como garantir que a "pasta" onde vamos guardar os arquivos esteja disponível.
    os.makedirs(diretorio_saida, exist_ok=True)

    # Percorre cada registro de clash (um registro de conflito entre objetos)
    for clash in clashs_total:
        # Extração dos dados relevantes do clash utilizando o método .get()
        # Caso a chave não exista, um valor vazio ('') é retornado.
        disp_1   = clash.get('disciplina_1', '')  # Primeira disciplina envolvida
        disp_2   = clash.get('disciplina_2', '')  # Segunda disciplina envolvida
        id_clash = clash.get('id', '')            # Identificador único do clash
        entity1  = clash.get('entity_1', '')      # Primeiro objeto (entidade) envolvido
        entity2  = clash.get('entity_2', '')      # Segundo objeto (entidade) envolvido
        layer1   = clash.get('layer_1', '')       # Layer (camada) correspondente ao primeiro objeto
        layer2   = clash.get('layer_2', '')       # Layer correspondente ao segundo objeto
        coord_x  = clash.get('coord_x', '')       # Coordenada X do clash
        coord_y  = clash.get('coord_y', '')       # Coordenada Y do clash
        coord_z  = clash.get('coord_z', '')       # Coordenada Z do clash

        # Criação do conteúdo que será escrito nos arquivos de texto.
        # Imagine isso como um "relatório" simples do clash, onde cada campo é exibido em uma linha.
        # O formato é definido para facilitar a leitura, similar a um bilhete com todas as informações importantes.
        # Exemplo de layout:
        #
        #   X: 100.0
        #   Y: 200.0
        #   Z: 50.0
        #   Objetos: Topografia X Geometria
        #   ID: 12345
        #   Entity1: 9876
        #   Entity2: 5432
        #   Layer1: L1
        #   Layer2: L2
        #   ----------------------------------------
        conteudo = f"X: {coord_x}\n" \
                   f"Y: {coord_y}\n" \
                   f"Z: {coord_z}\n" \
                   f"Objetos: {disp_1} X {disp_2}\n" \
                   f"ID: {id_clash}\n" \
                   f"Entity1: {entity1}\n" \
                   f"Entity2: {entity2}\n" \
                   f"Layer1: {layer1}\n" \
                   f"Layer2: {layer2}\n" \
                   f"{'-'*40}\n"

        # Lista as disciplinas envolvidas no clash, pois um mesmo clash pode estar 
        # relacionado a duas disciplinas (por exemplo, Topografia e Geometria).
        disciplinas = [disp_1, disp_2]

        # Para cada disciplina presente na lista:
        for disciplina in disciplinas:
            # Verifica se a disciplina não é uma string vazia
            if disciplina:
                # Utiliza o dicionário global 'clashs_contados' para rastrear se 
                # esse clash já foi contabilizado para a respectiva disciplina.
                # Isso serve para evitar escrita duplicada no mesmo arquivo,
                # como se você quisesse anotar apenas uma vez a ocorrência de um ocorrido.
                if id_clash not in clashs_contados[disciplina]:
                    # Define o caminho do arquivo de texto para aquela disciplina.
                    # Exemplo: se a disciplina for "Topografia", o caminho será "diretorio_saida/Topografia.txt"
                    caminho_txt = os.path.join(diretorio_saida, f"{disciplina}.txt")
                    
                    # Abre o arquivo em modo de adição (append 'a') para escrever o conteúdo.
                    # Se o arquivo já existir, o conteúdo será adicionado ao final, assim como atualizar uma lista.
                    with open(caminho_txt, 'a', encoding='utf-8') as txt_file:
                        txt_file.write(conteudo)
                    
                    # Registra que este clash (representado pelo seu id) já foi escrito para esta disciplina.
                    # Pense nisso como marcar uma tarefa como "feita" para que não seja repetida.
                    clashs_contados[disciplina].add(id_clash)

# ------------------------------------------------------------
# Função: criar_txt_defeitos
# Propósito: Cria um único arquivo 'defeitos.txt' com detalhes dos clashs
# que não possuem todas as informações obrigatórias.
# ------------------------------------------------------------
def criar_txt_defeitos(clashs_problematicos, diretorio_saida):
    os.makedirs(diretorio_saida, exist_ok=True)
    caminho_defeitos = os.path.join(diretorio_saida, "defeitos.txt")
    
    with open(caminho_defeitos, 'w', encoding='utf-8') as txt_file:
        for clash in clashs_problematicos:
            conteudo = f"Nome: {clash.get('name', 'N/A')}\n"
            conteudo += f"ID: {clash.get('id', 'N/A')}\n"
            conteudo += f"Coord X: {clash.get('coord_x', 'N/A')}\n"
            conteudo += f"Coord Y: {clash.get('coord_y', 'N/A')}\n"
            conteudo += f"Coord Z: {clash.get('coord_z', 'N/A')}\n"
            conteudo += f"Disciplina 1: {clash.get('disciplina_1', 'N/A')}\n"
            conteudo += f"Disciplina 2: {clash.get('disciplina_2', 'N/A')}\n"
            conteudo += f"Entity 1: {clash.get('entity_1', 'N/A')}\n"
            conteudo += f"Entity 2: {clash.get('entity_2', 'N/A')}\n"
            conteudo += f"Layer 1: {clash.get('layer_1', 'N/A')}\n"
            conteudo += f"Layer 2: {clash.get('layer_2', 'N/A')}\n"
            conteudo += "-"*40 + "\n"
            txt_file.write(conteudo)

# ------------------------------------------------------------
# Função: contagem_conflitos_totais
# Propósito: Iterar pelos clashs para determinar combinações únicas de layers
# e contar seus conflitos.
# ------------------------------------------------------------
def contagem_conflitos_totais(clashs):
    lista_conflitos = []
    contagem_conflitos_total = []
    for clash in clashs:
        if clash.get('layer_1') and clash.get('layer_2'):
            key = f"{clash['layer_1']}%{clash['layer_2']}"
            key_inv = f"{clash['layer_2']}%{clash['layer_1']}"
            if key not in lista_conflitos and key_inv not in lista_conflitos:
                lista_conflitos.append(key)
                contagem_conflitos_total.append(1)
            else:
                for i in range(len(lista_conflitos)):
                    if lista_conflitos[i] == key:
                        contagem_conflitos_total[i] += 1
    return lista_conflitos, contagem_conflitos_total

# ------------------------------------------------------------
# Função: separar_layers
# Propósito: Agrupar os layers de cada clash de acordo com suas disciplinas.
# Retorna um dicionário mapeando cada disciplina para a lista de layers associados.
# ------------------------------------------------------------
def separar_layers(clashs):
    disciplinas_layers = defaultdict(list)
    for clash in clashs:
        layers_disciplinas = []
        if 'disciplina_1' in clash and 'layer_1' in clash:
            layers_disciplinas.append((clash['layer_1'], clash['disciplina_1']))
        if 'disciplina_2' in clash and 'layer_2' in clash:
            layers_disciplinas.append((clash['layer_2'], clash['disciplina_2']))
        for layer, disciplina in layers_disciplinas:
            if layer not in disciplinas_layers[disciplina]:
                disciplinas_layers[disciplina].append(layer)
    return dict(disciplinas_layers)

# ------------------------------------------------------------
# Função: excel_conflitos_por_disciplina
# Propósito: Gerar uma planilha Excel listando os conflitos (layers e contagens)
# organizados por pares de disciplinas.
# ------------------------------------------------------------
def excel_conflitos_por_disciplina(conflitos_por_disciplina, dicionario_layer_disciplina, saida):
    # Cria uma nova planilha Excel e seleciona a primeira aba
    workbook = openpyxl.Workbook()
    aba = workbook.active

    # Variável que controla a numeração das linhas na planilha
    row_num = 1

    # Itera sobre cada par de disciplinas e os respectivos conflitos
    for disciplinas_chave, conflitos in conflitos_por_disciplina.items():
        # Adiciona os títulos das colunas para cada par de disciplinas
        aba.merge_cells(f'A{row_num}:B{row_num}')
        aba[f'A{row_num}'] = 'Disciplinas'
        aba[f'C{row_num}'] = 'Soma'
        row_num += 1

        # Separa as duas disciplinas usando o separador ' x '
        disciplina1, disciplina2 = disciplinas_chave.split(' x ')
        
        # Obtém o total de conflitos para esse par de disciplinas
        total_conflitos = conflitos['total']
        
        # Grava os nomes das disciplinas e o total de conflitos na primeira linha do grupo
        aba[f'A{row_num}'] = disciplina1      # Coluna A: Disciplina 1
        aba[f'B{row_num}'] = disciplina2      # Coluna B: Disciplina 2
        aba[f'C{row_num}'] = total_conflitos    # Coluna C: Total de conflitos
        row_num += 1

        # Se houver detalhes de conflitos (camadas específicas), insere um cabeçalho para os detalhes
        if conflitos['conflitos']:
            aba.merge_cells(f'A{row_num}:B{row_num}')
            aba[f'A{row_num}'] = 'Layers'  # Cabeçalho para Layer
            aba[f'C{row_num}'] = 'Contagem' # Cabeçalho para a contagem de conflitos
            row_num += 1

        # Itera sobre cada detalhe de conflito para esse par de disciplinas
        for layer_conflito, contagem in conflitos['conflitos'].items():
            # Cada chave 'layer_conflito' contém os dois layers separados por '%'
            layer1, layer2 = layer_conflito.split('%')
            
            # Verifica qual layer pertence à disciplina 1 para manter a consistência na ordenação
            if layer1 in dicionario_layer_disciplina.get(disciplina1, []):
                aba[f'A{row_num}'] = layer1  # Coloca layer1 na coluna A
                aba[f'B{row_num}'] = layer2  # Coloca layer2 na coluna B
            else:
                aba[f'A{row_num}'] = layer2  # Caso contrário, inverte a ordem
                aba[f'B{row_num}'] = layer1
            aba[f'C{row_num}'] = contagem  # Coloca a contagem de conflitos na coluna C
            row_num += 1  # Avança para a próxima linha

        # Adiciona uma linha em branco para separar os grupos de conflitos entre disciplinas
        row_num += 1

        # Estilização da planilha
        aba.column_dimensions['A'].width = 30
        aba.column_dimensions['B'].width = 30
        aba.column_dimensions['C'].width = 15

    # Salva a planilha no diretório de saída especificado com o nome 'lista_conflitos_disciplinas.xlsx'
    workbook.save(f'{saida}/lista_conflitos_disciplinas.xlsx')

# ------------------------------------------------------------
# Função: relacionar_conflitos_disciplinas
# Propósito: Relacionar os conflitos (camadas) entre pares de disciplinas,
# retornando um dicionário com a contagem total e os detalhes de cada par.
# ------------------------------------------------------------
def relacionar_conflitos_disciplinas(lista_conflitos, contagem_conflitos_total, lista_disciplinas, dicionario_layer_disciplina):
    """
    Relaciona os conflitos entre os layers de cada par de disciplinas, agrupando
    as contagens e os detalhes dos conflitos.

    Parâmetros:
      lista_conflitos (list): Lista contendo strings que representam os pares de layers em conflito,
                              onde os layers são separados por '%'.
      contagem_conflitos_total (list): Lista com a contagem (número de ocorrências) para cada
                                       combinação de layers em 'lista_conflitos'.
      lista_disciplinas (list): Lista de disciplinas encontradas durante o processamento.
      dicionario_layer_disciplina (dict): Dicionário que mapeia cada disciplina para a lista de layers
                                          associados a ela.
    
    Retorna:
      dict: Um dicionário que relaciona cada par de disciplinas (no formato "Disciplina1 x Disciplina2")
            com os detalhes dos conflitos (camadas conflitantes e suas contagens) e o total de conflitos.
    """
    conflitos_por_disciplina = {}
    
    # Alteramos o loop interno para iniciar em 'i' em vez de 'i + 1'
    # Assim, serão consideradas também as combinações onde ambas as disciplinas são iguais.
    for i in range(len(lista_disciplinas)):
        for j in range(i, len(lista_disciplinas)):  # Permite combinações iguais (ex: Drenagem x Drenagem) e evita duplicatas
            disciplina1 = lista_disciplinas[i]
            disciplina2 = lista_disciplinas[j]
            chave_disciplinas = f"{disciplina1} x {disciplina2}"

            conflitos_layer = {}
            total_conflitos = 0

            for k in range(len(lista_conflitos)):
                layer1, layer2 = lista_conflitos[k].split('%')
                contagem = contagem_conflitos_total[k]
                
                # Verificação: se ambos os layers estão associados às disciplinas correspondentes.
                # No caso em que disciplina1 e disciplina2 são iguais, a condição exige que
                # ambos os layers pertençam à mesma disciplina.
                if ((layer1 in dicionario_layer_disciplina.get(disciplina1, []) and
                     layer2 in dicionario_layer_disciplina.get(disciplina2, [])) or
                    (layer2 in dicionario_layer_disciplina.get(disciplina1, []) and
                     layer1 in dicionario_layer_disciplina.get(disciplina2, []))):
                    
                    conflitos_layer[lista_conflitos[k]] = contagem
                    total_conflitos += contagem

            conflitos_por_disciplina[chave_disciplinas] = {
                'conflitos': conflitos_layer,
                'total': total_conflitos
            }
    
    return conflitos_por_disciplina

# ------------------------------------------------------------
# Função: count_total_clashes
# Propósito: Retornar a quantidade total de clashs processados.
# ------------------------------------------------------------
def count_total_clashes(clashs):
    return len(clashs)

# ------------------------------------------------------------
# Função: process_files
# Propósito: Função principal que orquestra o processamento dos arquivos,
# atualiza a barra de progresso, gera os arquivos finais e exibe mensagens de status.
# ------------------------------------------------------------
def process_files():
    # Desabilita o botão de processamento para evitar múltiplos cliques
    # e reinicia a barra de progresso.
    process_button.config(state=tk.DISABLED)
    progress_bar['value'] = 0  
    progress_label.config(text="Processando arquivo de clash...")
    try:
        # **************************************************
        # Etapa 1: Processamento do arquivo de Clash
        # **************************************************
        clashs, lista_disciplinas, clashs_problematicos = process_clash_file(clash_file_path.get())
        
        # Atualiza a barra de progresso e o status para a próxima etapa.
        progress_bar['value'] = 25
        progress_label.config(text="Processando matriz...")
        
        # **************************************************
        # Etapa 2: Separação dos layers por disciplina
        # **************************************************
        dicionario_layers_por_disciplinas = separar_layers(clashs)
        
        # (DEBUG) - Descomente para visualizar o mapeamento disciplina/layers
        # print("=== RELAÇÃO DISCIPLINA/LAYERS ===")
        # for disciplina, layers in dicionario_layers_por_disciplinas.items():
        #     print(f"{disciplina}:")
        #     for layer in layers:
        #         print(f" - {layer}")
        
        # **************************************************
        # Etapa 3: Contagem e relacionamento dos conflitos entre layers
        # **************************************************
        clash_semi_aprovados = process_matrix(clashs, matrix_file_path.get())
        progress_bar['value'] = 50
        progress_label.config(text="Relacionando conflitos por disciplina...")
        
        # **************************************************
        # Etapa 4: Geração do relatório de conflitos por disciplina
        # **************************************************
        lista_conflitos, contagem_conflitos_total = contagem_conflitos_totais(clash_semi_aprovados)
        progress_bar['value'] = 75
        progress_label.config(text="Processando matriz de aprovação...")
        
        # Atualiza a lista de disciplinas baseada apenas nos clashs semi-aprovados.
        lista_disciplinas = []
        for clash in clash_semi_aprovados:
            disciplina1 = clash.get('disciplina_1')
            disciplina2 = clash.get('disciplina_2')
            if disciplina1 and disciplina1 not in lista_disciplinas:
                lista_disciplinas.append(disciplina1)
            if disciplina2 and disciplina2 not in lista_disciplinas:
                lista_disciplinas.append(disciplina2)

        # Agora, relaciona os conflitos considerando apenas os clashs semi-aprovados
        conflitos_por_disciplina = relacionar_conflitos_disciplinas(
            lista_conflitos, contagem_conflitos_total, lista_disciplinas, dicionario_layers_por_disciplinas
        )

        # Gera o arquivo Excel com os conflitos entre disciplinas para os clashs semi-aprovados
        excel_conflitos_por_disciplina(conflitos_por_disciplina, dicionario_layers_por_disciplinas, output_dir.get())
        
        # **************************************************
        # Etapa 5: Aprovação e verificação de exceções na matriz
        # **************************************************
        clashs_aprovados, clashs_excecoes = separacao_de_excecao(clash_semi_aprovados, matrix_file_path.get())
        
        # **************************************************
        # Etapa 6: Geração dos arquivos TXT por disciplina e dos defeitos
        # **************************************************
        criar_txts_por_disciplina(clashs_aprovados, output_dir.get())
        criar_txt_defeitos(clashs_problematicos, output_dir.get())
        
        # **************************************************
        # Etapa Final: Conclusão do processamento e exibição dos resultados
        # **************************************************
        total_clashes = count_total_clashes(clash_semi_aprovados)
        print("Clashes Totais:", total_clashes)
        print("Clashes Aprovados:", len(clashs_aprovados))
        print("Clashes Exceções:", len(clashs_excecoes))
        
        progress_bar['value'] = 100
        progress_label.config(text="Concluído!")
        messagebox.showinfo("Sucesso", "Processamento concluído com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro durante o processamento: {str(e)}")
        progress_label.config(text="Erro durante o processamento.")
    finally:
        process_button.config(state=tk.NORMAL)

# ------------------------------------------------------------
# Função: start_processing
# Propósito: Inicia o processamento dos arquivos em uma thread separada,
# evitando travamentos na interface gráfica.
# ------------------------------------------------------------
def start_processing():
    thread = threading.Thread(target=process_files)
    thread.start()

# ------------------------------------------------------------
# Função: create_gui
# Propósito: Monta a interface gráfica utilizando Tkinter, configurando
# os campos de seleção de arquivos, barra de progresso e botões.
# ------------------------------------------------------------
def create_gui():
    global clash_file_path, matrix_file_path, output_dir, process_button, progress_bar, progress_label
    root = tk.Tk()
    root.title("Clash Analyzer")
    root.geometry("600x450")
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    
    # Configuração do ícone
    if getattr(sys, 'frozen', False):
        icon_path = os.path.join(sys._MEIPASS, 'icon.ico')
    else:
        icon_path = 'icon.ico'
    try:
        root.iconbitmap(icon_path)
    except:
        pass
    
    # Inicializa as variáveis globais como StringVar's
    clash_file_path = tk.StringVar()
    matrix_file_path = tk.StringVar()
    output_dir = tk.StringVar()
    
    # Criação do frame principal utilizando ttk para um visual mais moderno
    main_frame = ttk.Frame(root, padding="20")
    main_frame.grid(column=0, row=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    # Configuração dos widgets (Labels, Entries e Botões) com grid para melhor alinhamento
    ttk.Label(main_frame, text="Arquivo de Clash:").grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)
    clash_entry = ttk.Entry(main_frame, textvariable=clash_file_path, width=50)
    clash_entry.grid(column=1, row=0, sticky=(tk.W, tk.E), padx=5, pady=5)
    ttk.Button(main_frame, text="Selecionar Arquivo", command=select_clash_file).grid(column=2, row=0, sticky=tk.W, padx=5, pady=5)

    ttk.Label(main_frame, text="Arquivo da Matriz:").grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)
    matrix_entry = ttk.Entry(main_frame, textvariable=matrix_file_path, width=50)
    matrix_entry.grid(column=1, row=1, sticky=(tk.W, tk.E), padx=5, pady=5)
    ttk.Button(main_frame, text="Selecionar Matriz", command=select_matrix_file).grid(column=2, row=1, sticky=tk.W, padx=5, pady=5)

    ttk.Label(main_frame, text="Diretório de Saída:").grid(column=0, row=2, sticky=tk.W, padx=5, pady=5)
    output_entry = ttk.Entry(main_frame, textvariable=output_dir, width=50)
    output_entry.grid(column=1, row=2, sticky=(tk.W, tk.E), padx=5, pady=5)
    ttk.Button(main_frame, text="Selecionar Diretório", command=select_output_dir).grid(column=2, row=2, sticky=tk.W, padx=5, pady=5)
    
    # Barra de Progresso e Label de Status
    progress_bar = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, length=300, mode='determinate')
    progress_bar.grid(column=0, row=3, columnspan=3, sticky=(tk.W, tk.E), padx=5, pady=10)
    progress_label = ttk.Label(main_frame, text="")
    progress_label.grid(column=0, row=4, columnspan=3, sticky=tk.W, padx=5)
    
    # Botão de Processamento
    process_button = ttk.Button(main_frame, text="Processar", command=start_processing)
    process_button.grid(column=1, row=5, pady=20)
    
    root.mainloop()

if __name__ == "__main__":
    create_gui()